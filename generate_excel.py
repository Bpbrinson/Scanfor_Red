"""
generate_excel.py — Excel Report Generator for Scanfor_Red
===========================================================
Reads a Scanfor_Red JSON output file and writes a formatted
Excel workbook to outputs/json_to_excel/<stem>/<stem>.xlsx

Called automatically by scanfor_red.py after each scan.
Can also be run standalone:
    python generate_excel.py Dashboard_Screenshot/6_4.json
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("[ERROR] openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

# Optional enrichment — applied to records that pre-date the enrich.py integration
try:
    from enrich import load_service_lookup, enrich_with_service_info as _enrich
    _LOOKUP = load_service_lookup()
    def _ensure_enriched(records):
        return [_enrich(r, _LOOKUP) if not r.get("service_name") else r
                for r in records]
except Exception:
    def _ensure_enriched(records):
        return records


# ── Colour palette ────────────────────────────────────────────────────────────
NAVY        = "1F3864"
BLUE        = "305496"
LIGHT_BLUE  = "D6E4F0"
MID_BLUE    = "BDD7EE"
GREEN_HDR   = "375623"
GREEN_FILL  = "E2EFDA"
AMBER_HDR   = "7F4F0E"
AMBER_FILL  = "FFF2CC"
RED_FILL    = "FCE4EC"
WHITE       = "FFFFFF"
OFF_WHITE   = "F2F7FC"
GRAY_BORDER = "BFBFBF"
DARK_TEXT   = "000000"
WHITE_TEXT  = "FFFFFF"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=DARK_TEXT, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _border():
    thin = Side(style="thin", color=GRAY_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header_row(ws, row, col_start, col_end, fill_hex, text_color=WHITE_TEXT):
    """Apply header formatting across a row range."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True, color=text_color)
        cell.border = _border()
        cell.alignment = _center()


def _style_data_row(ws, row, col_start, col_end, alt=False):
    fill_hex = OFF_WHITE if alt else WHITE
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(fill_hex)
        cell.border = _border()
        cell.alignment = Alignment(vertical="top", wrap_text=False)


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _set(ws, row, col, value, bold=False, fill_hex=None, color=DARK_TEXT,
         align=None, size=11, italic=False, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color, size=size, italic=italic)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if align:
        cell.alignment = align
    if num_format:
        cell.number_format = num_format
    cell.border = _border()
    return cell


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_data_sheet(ws, records):
    """Full alert table with enrichment columns (reorganized layout)."""
    headers = [
        "System", "Log File", "Col", "Count",
        "Service", "Error Description", "Remediation Hint", "Working On",
        "Ticket Exists", "Ticket ID", "Ticket Notes",
    ]
    col_widths = [42, 62, 7, 9, 18, 32, 52, 52, 26, 14, 16]

    # Header row
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _set(ws, 1, c, h, bold=True, fill_hex=BLUE, color=WHITE_TEXT,
             align=_center())
        _col_width(ws, c, w)

    # Data rows
    for r_idx, rec in enumerate(records, start=2):
        alt = (r_idx % 2 == 0)
        _style_data_row(ws, r_idx, 1, len(headers), alt=alt)
        em = rec.get("error_meaning", {})
        row_vals = [
            rec.get("system", ""),
            rec.get("log_file", ""),
            rec.get("column"),
            rec.get("count"),
            rec.get("service_name", ""),
            em.get("error_description", ""),
            em.get("remediation_hint", ""),
            em.get("escalation_team", ""),   # "Working On"
            rec.get("ticket_exists", ""),    # Ticket Exists — from the web frontend
            rec.get("ticket_id", ""),        # Ticket ID — from the web frontend
            rec.get("ticket_notes", ""),     # Ticket Notes — from the web frontend
        ]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.alignment = _wrap() if c in (3, 4, 5) else Alignment(
                vertical="top", wrap_text=(c >= 6)
            )
            # Highlight manual ticket columns with amber tint
            if c in (9, 10, 11):
                cell.fill = _fill(AMBER_FILL)
                cell.font = _font(color=AMBER_HDR, italic=True)

    # Freeze header, set row heights
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20
    for r_idx in range(2, len(records) + 2):
        ws.row_dimensions[r_idx].height = 36

    # Excel table
    if records:
        last_row = len(records) + 1
        tbl = Table(displayName="AlertData",
                    ref=f"A1:{get_column_letter(len(headers))}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(tbl)


def _build_summary_sheet(ws, records, source_file):
    """KPI summary and top-10 systems."""
    systems   = sorted({r["system"] for r in records})
    services  = sorted({r.get("service_name", "") for r in records if r.get("service_name")})
    log_files = sorted({r["log_file"] for r in records})
    total     = sum((r.get("count") or 0) for r in records)
    peak      = max((r.get("count") or 0) for r in records) if records else 0

    # Title banner
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Scanfor Red — Daily Alert Summary"
    title.font  = _font(bold=True, color=WHITE_TEXT, size=14)
    title.fill  = _fill(NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # KPI block
    kpi_headers = ["Metric", "Value"]
    kpis = [
        ("Source file",       source_file),
        ("Generated",         datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total red alerts",  len(records)),
        ("Unique systems",    len(systems)),
        ("Unique services",   len(services)),
        ("Unique log files",  len(log_files)),
        ("Total error count", total),
        ("Peak single cell",  peak),
    ]
    for c, h in enumerate(kpi_headers, start=1):
        _set(ws, 3, c, h, bold=True, fill_hex=BLUE, color=WHITE_TEXT, align=_center())
    for r_idx, (label, val) in enumerate(kpis, start=4):
        alt = (r_idx % 2 == 0)
        _set(ws, r_idx, 1, label, fill_hex=LIGHT_BLUE if not alt else OFF_WHITE)
        cell = _set(ws, r_idx, 2, val,
                    fill_hex=WHITE if alt else OFF_WHITE,
                    num_format="#,##0" if isinstance(val, int) else None)
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Top-10 systems by total count
    sys_totals = sorted(
        [{"system": s,
          "alerts": sum(1 for r in records if r["system"] == s),
          "total":  sum((r.get("count") or 0) for r in records if r["system"] == s)}
         for s in systems],
        key=lambda x: x["total"], reverse=True
    )[:10]

    top_row = 3
    for c, h in enumerate(["System", "Alerts", "Total Count"], start=4):
        _set(ws, top_row, c, h, bold=True, fill_hex=GREEN_HDR, color=WHITE_TEXT,
             align=_center())
    for r_idx, row in enumerate(sys_totals, start=top_row + 1):
        alt = (r_idx % 2 == 0)
        fill = OFF_WHITE if alt else WHITE
        _set(ws, r_idx, 4, row["system"],    fill_hex=fill)
        _set(ws, r_idx, 5, row["alerts"],    fill_hex=fill, num_format="#,##0",
             align=_center())
        _set(ws, r_idx, 6, row["total"],     fill_hex=fill, num_format="#,##0",
             align=_center())

    # Column widths
    for col, w in [(1, 26), (2, 22), (3, 4), (4, 46), (5, 12), (6, 14)]:
        _col_width(ws, col, w)

    ws.sheet_view.showGridLines = False


def _build_by_system_sheet(ws, records):
    """One row per system — alert count, total errors, services seen."""
    systems = sorted({r["system"] for r in records})
    headers = ["System", "Services", "Alert Rows", "Total Count"]
    col_widths = [46, 36, 36, 13]

    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _set(ws, 1, c, h, bold=True, fill_hex=BLUE, color=WHITE_TEXT, align=_center())
        _col_width(ws, c, w)

    rows = []
    for sys_name in systems:
        recs = [r for r in records if r["system"] == sys_name]
        services = ", ".join(sorted({r.get("service_name", "") for r in recs if r.get("service_name")}))
        rows.append([sys_name, services, len(recs),
                     sum((r.get("count") or 0) for r in recs)])

    rows.sort(key=lambda x: x[3], reverse=True)

    for r_idx, row_vals in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        _style_data_row(ws, r_idx, 1, len(headers), alt=alt)
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c == 2),
                                       horizontal="right" if c >= 3 else "left")
            if c >= 3:
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    if rows:
        tbl = Table(displayName="BySystem",
                    ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
        ws.add_table(tbl)


def _build_by_column_sheet(ws, records):
    """One row per column number — maps column index to error type."""
    ERROR_PATTERNS = {
        0:  "^20..-..-..[T ]..:..:..\\.\\d+ Error",
        1:  "conform.to.algorithm.constraints",
        2:  "java[.].*exception",
        3:  "java.lang.NullPointerException",
        4:  "java.jdbc",
        5:  "javax.net.ssl",
        6:  "Bind.Failed",
        7:  "^Exception",
        8:  "java.lang.NoClassDefFoundError",
        9:  "Algorithm.negotiation.fail",
        10: "ExpiredToken",
        11: "SQLException",
        12: "credentials",
        13: "Return.Code..408",
    }
    cols_seen = sorted({r["column"] for r in records if r.get("column") is not None})
    headers = ["Column", "Error Pattern", "Alert Rows", "Total Count", "Systems Affected"]
    col_widths = [9, 46, 13, 14, 46]

    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _set(ws, 1, c, h, bold=True, fill_hex=BLUE, color=WHITE_TEXT, align=_center())
        _col_width(ws, c, w)

    rows = []
    for col_idx in cols_seen:
        recs = [r for r in records if r.get("column") == col_idx]
        systems = ", ".join(sorted({r["system"] for r in recs}))
        rows.append([col_idx,
                     ERROR_PATTERNS.get(col_idx, f"Column {col_idx}"),
                     len(recs),
                     sum((r.get("count") or 0) for r in recs),
                     systems])

    rows.sort(key=lambda x: x[3], reverse=True)

    for r_idx, row_vals in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        _style_data_row(ws, r_idx, 1, len(headers), alt=alt)
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c in (2, 5)),
                                       horizontal="center" if c in (1, 3, 4) else "left")
            if c in (3, 4):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    if rows:
        tbl = Table(displayName="ByColumn",
                    ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
        ws.add_table(tbl)


def _build_ticket_sheet(ws, records):
    """Pre-populated ticket review sheet — amber columns for manual entry."""
    # One row per unique system+service combination
    combos = {}
    for r in records:
        key = (r.get("system", ""), r.get("service_name", ""))
        if key not in combos:
            combos[key] = {"count": 0, "total": 0, "columns": [],
                           "tickets": [], "ids": [], "notes": [],
                           "friendly": r.get("error_meaning", {}).get("friendly_name", ""),
                           "team": r.get("error_meaning", {}).get("escalation_team", "")}
        combos[key]["count"] += 1
        combos[key]["total"] += (r.get("count") or 0)
        col = r.get("column")
        if col is not None and col not in combos[key]["columns"]:
            combos[key]["columns"].append(col)
        for field, bucket in (("ticket_exists", "tickets"),
                              ("ticket_id", "ids"),
                              ("ticket_notes", "notes")):
            v = (r.get(field) or "").strip()
            if v and v not in combos[key][bucket]:
                combos[key][bucket].append(v)

    headers = [
        "System", "Service", "Friendly Name",
        "Alert Rows", "Total Count", "Escalation Team",
        "Column", "Error Friendly Name",
        "Ticket Exists?", "Ticket ID", "Ticket Notes",
    ]
    col_widths = [42, 20, 32, 13, 14, 26, 15, 16, 36, 13, 13]
    MANUAL_COLS = {8, 9, 10, 11}   # Error Friendly Name + 3 ticket columns

    # Title (merged across A1:I1 to match the template)
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "Daily Ticket Review"
    title.font  = _font(bold=True, color=WHITE_TEXT, size=12)
    title.fill  = _fill(AMBER_HDR)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        fill = AMBER_FILL if c in MANUAL_COLS else BLUE
        text = AMBER_HDR if c in MANUAL_COLS else WHITE_TEXT
        _set(ws, 2, c, h, bold=True, fill_hex=fill, color=text, align=_center())
        _col_width(ws, c, w)

    rows = sorted(combos.items(), key=lambda x: x[1]["total"], reverse=True)
    for r_idx, ((sys_name, svc), data) in enumerate(rows, start=3):
        alt = (r_idx % 2 == 0)
        base_fill = OFF_WHITE if alt else WHITE
        for c in range(1, len(headers) + 1):
            fill = AMBER_FILL if c in MANUAL_COLS else base_fill
            cell = ws.cell(row=r_idx, column=c)
            cell.fill  = _fill(fill)
            cell.border = _border()
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (3, 11)),
                                       horizontal="center" if c in (4, 5, 7) else "left")
            if c in MANUAL_COLS:
                cell.font = _font(italic=True, color=AMBER_HDR)

        col_str = ", ".join(str(x) for x in data["columns"])
        # Aggregated manual ticket fields for this system+service combo.
        ticket_str = ", ".join(data["tickets"])
        id_str = ", ".join(data["ids"])
        notes_str = ", ".join(data["notes"])
        vals = [sys_name, svc, data["friendly"], data["count"],
                data["total"], data["team"], col_str, "",
                ticket_str, id_str, notes_str]
        for c, val in enumerate(vals, start=1):
            if val != "":
                ws.cell(row=r_idx, column=c, value=val)
                if c in (4, 5):
                    ws.cell(row=r_idx, column=c).number_format = "#,##0"

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_excel(json_path: str, output_dir: str) -> str:
    """
    Read a Scanfor_Red JSON file and write a formatted Excel workbook.

    Args:
        json_path:  Path to the .json output from scanfor_red.py
        output_dir: Folder to create and write the .xlsx into

    Returns:
        Absolute path of the created .xlsx file
    """
    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    if not isinstance(records, list):
        raise ValueError(f"Expected a JSON array in {json_path}")

    records = _ensure_enriched(records)

    stem = os.path.splitext(os.path.basename(json_path))[0]
    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = os.path.join(output_dir, f"{stem}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)   # remove the default blank sheet

    # Sheet order
    ws_summary = wb.create_sheet("Summary")
    ws_data     = wb.create_sheet("Data")
    ws_system   = wb.create_sheet("By System")
    ws_column   = wb.create_sheet("By Column")
    ws_ticket   = wb.create_sheet("Ticket Check")

    # Tab colours
    ws_summary.sheet_properties.tabColor = NAVY
    ws_data.sheet_properties.tabColor    = BLUE
    ws_system.sheet_properties.tabColor  = GREEN_HDR
    ws_column.sheet_properties.tabColor  = "7B7B7B"
    ws_ticket.sheet_properties.tabColor  = AMBER_HDR

    for ws in [ws_summary, ws_data, ws_system, ws_column, ws_ticket]:
        ws.sheet_view.showGridLines = False

    _build_summary_sheet(ws_summary, records, os.path.basename(json_path))
    _build_data_sheet(ws_data, records)
    _build_by_system_sheet(ws_system, records)
    _build_by_column_sheet(ws_column, records)
    _build_ticket_sheet(ws_ticket, records)

    wb.save(xlsx_path)
    return os.path.abspath(xlsx_path)


# ── Standalone CLI ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <path/to/alerts.json> [output_dir]")
        sys.exit(1)

    json_file  = sys.argv[1]
    stem       = os.path.splitext(os.path.basename(json_file))[0]
    out_dir    = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.abspath(json_file)),
        "..", "outputs", "json_to_excel", stem
    )
    out_dir = os.path.normpath(out_dir)

    path = generate_excel(json_file, out_dir)
    print(f"Excel written -> {path}")
